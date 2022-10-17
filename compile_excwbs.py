#personal module
import df_metadata

import os
import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows

def scan_excel_files(folder_path):
    folder_path = str(folder_path)
    excel_files = []
    with os.scandir(folder_path) as folder:
        for files in folder:
            f = files.name
            period_pos = f.find('.')
            ext = f[period_pos:(len(f)+1)]
            if ext == '.xlsx':
                excel_files.append(files)
    
    return excel_files

def create_datadict_wb(search_folder_path, save_path, include_datadict=True):

    excels = scan_excel_files(search_folder_path)

    rows = []

    for f in excels:
        fname = f.name
        fpath = f.path
        #get sheet names for each excel wb
        wb_temp = xl.load_workbook(fpath, read_only=True)
        sheets = wb_temp.sheetnames
        for s in sheets:
            d = pd.read_excel(fpath, sheet_name=s)
            for col in d:
                r = {'file_name':fname, 'sheet_name':s, 'column_name':col, 'file_path':fpath}
                rows.append(r)
    
    #create list of scanned excel file names, their sheets, and their paths
    df = pd.DataFrame(rows)

    #add scanned excel files to first sheet
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'scanned files'
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    
    #add data dict to new sheet in wb if specified
    if include_datadict == True:
        fpaths = df['file_path'].unique()

        df_meta = df.set_index(['file_name', 'sheet_name', 'column_name'])
        for file in fpaths:
            fname = df.query(df['file_path == @file'])
            fname = fname['file_name'].unique().iat[0,0]
            #df_filtered = df_dd.query('file_path == @file')
            dict_sheetdata = pd.read_excel(file, sheet_name=None)
            for s, df in dict_sheetdata.items():
                df_updateidx = df
                df_updateidx['file_name'] = fname
                df_updateidx['sheet_name'] = s
                df_updateidx['column_name'] = 
            
            
            chk = dict_sheetdata['Trans']
            print(type(chk))
            print(chk)
            print('stop')

            
            #pass dict containing dfs of all sheets in file to metadata functions

#.set_index(['file_name', 'sheet_name', 'column_name'])

    wb.save(save_path)

    return wb


search_fpath = f''
save_path = f''

wb = create_datadict_wb(search_fpath, save_path)

wb.save(save_path)

df_files = pd.read_excel(save_path, sheet_name='scanned files')
print(df_files)

